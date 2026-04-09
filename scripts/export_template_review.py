"""Export just the task template for review."""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='1B1464')
PHASE_FILL = PatternFill('solid', fgColor='F3F0FF')
BORDER = Border(bottom=Side(style='thin', color='E7E5E4'), right=Side(style='thin', color='E7E5E4'))
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='top')
YELLOW = PatternFill('solid', fgColor='FFF9C4')

wb = Workbook()
ws = wb.active
ws.title = 'Task Template Review'

headers = [
    '#', 'Task Name', 'Phase', 'Owner', 'Lead Time\n(Business Days)',
    'Depends On (must finish before this starts)',
    'Notes',
    'CORRECT\nLead Time?', 'CORRECT\nDependencies?', 'CORRECT\nOwner?',
    'REMOVE\nthis task?', 'Other Notes / Changes'
]
widths = [5, 44, 24, 16, 14, 52, 44, 14, 30, 14, 12, 44]

from openpyxl.utils import get_column_letter
for col, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 40
ws.freeze_panes = 'A2'

# Parse template
template_file = '/Users/juliaford/Desktop/ee-gtm-planner/src/lib/task-template.ts'
with open(template_file) as f:
    content = f.read()

entries = []
blocks = content.split("name: '")[1:]
for block in blocks:
    name = block.split("'")[0]
    lt = int(m.group(1)) if (m := re.search(r'leadTime:\s*(\d+)', block)) else 0
    deps_match = re.search(r"dependsOn:\s*\[(.*?)\]", block, re.DOTALL)
    deps = ', '.join(re.findall(r"'([^']+)'", deps_match.group(1))) if deps_match else ''
    owner = m.group(1) if (m := re.search(r"owner:\s*'([^']+)'", block)) else ''
    phase = m.group(1) if (m := re.search(r"phase:\s*'([^']+)'", block)) else ''
    notes = m.group(1) if (m := re.search(r"notes:\s*'([^']*)'", block)) else ''
    is_optional = bool(re.search(r'isOptional:\s*true', block))
    is_manual = bool(re.search(r'isManualDate:\s*true', block))
    is_meeting = bool(re.search(r'isMeeting:\s*true', block))

    prefix = ''
    if is_optional: prefix = '[OPTIONAL] '
    if is_manual: prefix = '[MANUAL DATE] '
    if is_meeting: prefix = '[MEETING] '

    phase_labels = {
        'content_planning': '1 — Content Planning',
        'finalize_mgmt': '2 — Finalize & Inform Mgmt',
        'content_production': '3 — Content Production',
        'design_production': '4 — Design Production',
        'packaging': '2b — Packaging',
    }
    owner_labels = {
        'marketing': 'Marketing', 'creative': 'Creative', 'product': 'Growth',
        'influencer': 'Influencer', 'social': 'Social', 'ops': 'Operations',
        'copywriter': 'Copywriter', 'growth': 'Growth', 'leadership': 'Leadership',
        'channel_leads': 'Channel Leads', 'retail': 'Retail', 'pr': 'PR',
        'digital': 'Digital', 'external': 'External',
    }

    entries.append({
        'name': prefix + name,
        'phase': phase_labels.get(phase, phase),
        'owner': owner_labels.get(owner, owner),
        'lead_time': lt,
        'deps': deps,
        'notes': notes,
        'raw_phase': phase,
    })

# Write rows, grouped by phase with phase header rows
current_phase = ''
row = 2
for i, e in enumerate(entries):
    if e['phase'] != current_phase:
        current_phase = e['phase']
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PHASE_FILL
            cell.border = BORDER
        ws.cell(row=row, column=2, value=current_phase).font = Font(name='Arial', bold=True, size=11, color='1B1464')
        ws.row_dimensions[row].height = 28
        row += 1

    values = [i + 1, e['name'], e['phase'], e['owner'], e['lead_time'], e['deps'], e['notes'],
              '', '', '', '', '']
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.border = BORDER
        if col == 1: cell.alignment = CENTER
        elif col == 5: cell.alignment = CENTER
        elif col in (2, 6, 7, 12): cell.alignment = WRAP
        else: cell.alignment = Alignment(vertical='top')

    # Yellow review columns
    for col in (8, 9, 10, 11, 12):
        ws.cell(row=row, column=col).fill = YELLOW

    row += 1

# Instructions tab
ws2 = wb.create_sheet('Instructions')
ws2.column_dimensions['A'].width = 80
instructions = [
    "HOW TO REVIEW THIS TEMPLATE",
    "",
    "This is the master task template that generates ALL launch schedules.",
    "Every launch gets these same tasks, scheduled backward from the launch date.",
    "",
    "WHAT TO CHECK:",
    "  1. Are the tasks correct? Should any be removed or added?",
    "  2. Are the lead times (business days) right?",
    "  3. Are the dependencies right? (Task X must finish before Task Y starts)",
    "  4. Are the owners right?",
    "",
    "HOW TO MARK CHANGES:",
    "  - Yellow columns H-L are for your notes",
    "  - Column H: If lead time is wrong, write the correct number",
    "  - Column I: If dependencies are wrong, write what they should be",
    "  - Column J: If owner is wrong, write the correct owner",
    "  - Column K: Write 'YES' to remove a task",
    "  - Column L: Any other changes (rename, add new tasks, etc.)",
    "",
    "AFTER REVIEW:",
    "  Share this file back and all changes will be applied at once.",
    "  All 12 launches will be regenerated from the corrected template.",
]
for i, line in enumerate(instructions, 1):
    cell = ws2.cell(row=i, column=1, value=line)
    cell.font = Font(name='Arial', size=12 if i == 1 else 11, bold=(i == 1))
    cell.alignment = Alignment(wrap_text=True)

output = '/Users/juliaford/Desktop/GTM_Template_Review.xlsx'
wb.save(output)
print(f"Exported to: {output}")
print(f"  {len(entries)} tasks across {len(set(e['raw_phase'] for e in entries))} phases")
print(f"  Yellow columns are for your review — fill in corrections!")
