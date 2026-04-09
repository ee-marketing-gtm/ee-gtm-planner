"""
Export all launch task data from the Turso database to a review spreadsheet.
Two tabs:
  1. All Tasks — every task across every launch with dates, deps, owner, status
  2. Task Template — the master template with lead times and dependencies
"""
import json, subprocess, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Fetch data directly from Turso DB ───────────────────────────────
import urllib.request, os
from pathlib import Path

# Load env vars from .env.local
env_path = Path(__file__).parent.parent / '.env.local'
env_vars = {}
if env_path.exists():
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            k, v = line.split('=', 1)
            env_vars[k.strip()] = v.strip()

db_url = env_vars.get('TURSO_DATABASE_URL', os.environ.get('TURSO_DATABASE_URL', ''))
db_token = env_vars.get('TURSO_AUTH_TOKEN', os.environ.get('TURSO_AUTH_TOKEN', ''))

if not db_url:
    print("No TURSO_DATABASE_URL found in .env.local")
    sys.exit(1)

# Convert libsql:// to https:// for HTTP API
http_url = db_url.replace('libsql://', 'https://') + '/v2/pipeline'

try:
    req_body = json.dumps({
        "requests": [
            {"type": "execute", "stmt": {"sql": "SELECT value FROM kv_store WHERE key = 'launches'"}},
            {"type": "close"}
        ]
    }).encode()
    req = urllib.request.Request(http_url, data=req_body, headers={
        'Authorization': f'Bearer {db_token}',
        'Content-Type': 'application/json',
    })
    with urllib.request.urlopen(req, timeout=10) as resp:
        resp_data = json.loads(resp.read())

    rows = resp_data['results'][0]['response']['result']['rows']
    if rows:
        launches = json.loads(rows[0][0]['value'])
    else:
        launches = []
except Exception as e:
    print(f"Could not fetch from Turso DB: {e}")
    sys.exit(1)

if not launches:
    print("No launches found in database.")
    sys.exit(1)

print(f"Found {len(launches)} launches")

# ── Styles ──────────────────────────────────────────────────────────
HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='1B1464')
SUBHEADER_FILL = PatternFill('solid', fgColor='E8E6F0')
OVERDUE_FILL = PatternFill('solid', fgColor='FEE2E2')
COMPLETE_FILL = PatternFill('solid', fgColor='DCFCE7')
LAUNCH_FILL = PatternFill('solid', fgColor='FFF0F7')
BORDER = Border(
    bottom=Side(style='thin', color='E7E5E4'),
    right=Side(style='thin', color='E7E5E4'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='top')

wb = Workbook()

# ════════════════════════════════════════════════════════════════════
# TAB 1: ALL TASKS (grouped by launch)
# ════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'All Tasks'

headers = [
    'Launch', 'Launch Date', 'Sephora Date', 'Status (Launch)',
    'Task #', 'Task Name', 'Phase', 'Owner', 'Due Date', 'Start Date',
    'Duration (BD)', 'Dependencies', 'Task Status',
    'Notes', 'REVIEW — Flag Issues Here'
]
col_widths = [28, 14, 14, 14, 8, 42, 22, 16, 14, 14, 12, 42, 16, 40, 40]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

row = 2
from datetime import datetime, date
today = date.today().isoformat()

for launch in sorted(launches, key=lambda l: l.get('launchDate', '')):
    if launch.get('status') == 'archived':
        continue

    # Launch header row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = LAUNCH_FILL
        cell.border = BORDER

    ws.cell(row=row, column=1, value=launch['name']).font = Font(name='Arial', bold=True, size=11)
    ws.cell(row=row, column=2, value=launch.get('launchDate', ''))
    ws.cell(row=row, column=3, value=launch.get('sephoraLaunchDate', ''))
    ws.cell(row=row, column=4, value=launch.get('status', ''))
    row += 1

    tasks = sorted(launch.get('tasks', []), key=lambda t: t.get('sortOrder', 999))
    for i, task in enumerate(tasks, 1):
        due = task.get('dueDate', '')
        status = task.get('status', 'not_started')
        is_overdue = due and due < today and status not in ('complete', 'skipped')

        phase_map = {
            'content_planning': 'Content Planning',
            'finalize_mgmt': 'Finalize & Inform Mgmt',
            'content_production': 'Content Production',
            'design_production': 'Design Production',
            'packaging': 'Packaging',
        }
        owner_map = {
            'marketing': 'Marketing', 'channel_leads': 'Channel Leads',
            'creative': 'Creative', 'product': 'Growth',
            'retail': 'Retail', 'influencer': 'Influencer',
            'pr': 'PR', 'digital': 'Digital', 'ops': 'Operations',
            'social': 'Social', 'external': 'Copywriter',
        }
        status_map = {
            'not_started': 'Not Started', 'in_progress': 'In Progress',
            'complete': 'Complete', 'blocked': 'Stuck',
            'skipped': 'Skipped', 'waiting_review': 'Waiting Review',
        }

        deps = ', '.join(task.get('dependencyNames', []) or [])
        values = [
            '', '', '', '',
            i,
            task.get('name', ''),
            phase_map.get(task.get('phase', ''), task.get('phase', '')),
            owner_map.get(task.get('owner', ''), task.get('owner', '')),
            due,
            task.get('startDate', ''),
            task.get('durationDays', ''),
            deps,
            status_map.get(status, status),
            task.get('notes', ''),
            '',  # Review column — empty for Julia to fill
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = WRAP if col_idx in (6, 12, 14, 15) else CENTER if col_idx in (5, 9, 10, 11, 13) else Alignment(vertical='top')
            cell.border = BORDER
            cell.font = Font(name='Arial', size=10)

        # Highlight overdue rows
        if is_overdue:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = OVERDUE_FILL
        elif status == 'complete':
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = COMPLETE_FILL

        # Yellow the review column
        review_cell = ws.cell(row=row, column=15)
        review_cell.fill = PatternFill('solid', fgColor='FFF9C4')

        row += 1

    row += 1  # blank row between launches

# ════════════════════════════════════════════════════════════════════
# TAB 2: MASTER TEMPLATE
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Task Template')

tmpl_headers = [
    'Task Name', 'Phase', 'Owner', 'Lead Time (BD)',
    'Depends On', 'Notes',
    'REVIEW — Correct Lead Time', 'REVIEW — Correct Dependencies', 'REVIEW — Other Notes'
]
tmpl_widths = [42, 22, 16, 14, 50, 50, 18, 40, 40]

for col_idx, (header, width) in enumerate(zip(tmpl_headers, tmpl_widths), 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

ws2.freeze_panes = 'A2'

# Read template data from the TypeScript source
import re
template_file = '/Users/juliaford/Desktop/ee-gtm-planner/src/lib/task-template.ts'
with open(template_file, 'r') as f:
    content = f.read()

# Parse template entries
tasks_raw = re.findall(r'\{[^}]+name:\s*[\'"]([^\'"]+)[\'"][^}]*\}', content, re.DOTALL)

# More robust parsing
template_entries = []
blocks = content.split("name: '")[1:]
for block in blocks:
    name = block.split("'")[0]
    lt_match = re.search(r'leadTime:\s*(\d+)', block)
    lead_time = int(lt_match.group(1)) if lt_match else 0

    deps_match = re.search(r"dependsOn:\s*\[(.*?)\]", block, re.DOTALL)
    deps = ''
    if deps_match:
        deps = ', '.join(re.findall(r"'([^']+)'", deps_match.group(1)))

    owner_match = re.search(r"owner:\s*'([^']+)'", block)
    owner = owner_match.group(1) if owner_match else ''

    phase_match = re.search(r"phase:\s*'([^']+)'", block)
    phase = phase_match.group(1) if phase_match else ''

    notes_match = re.search(r"notes:\s*'([^']*)'", block)
    notes = notes_match.group(1) if notes_match else ''

    optional_match = re.search(r'isOptional:\s*true', block)
    manual_match = re.search(r'isManualDate:\s*true', block)

    phase_labels = {
        'content_planning': 'Content Planning',
        'finalize_mgmt': 'Finalize & Inform Mgmt',
        'content_production': 'Content Production',
        'design_production': 'Design Production',
        'packaging': 'Packaging',
    }
    owner_labels = {
        'marketing': 'Marketing', 'creative': 'Creative', 'product': 'Growth',
        'influencer': 'Influencer', 'social': 'Social', 'ops': 'Operations',
        'copywriter': 'Copywriter', 'growth': 'Growth', 'leadership': 'Leadership',
        'channel_leads': 'Channel Leads', 'retail': 'Retail', 'pr': 'PR',
        'digital': 'Digital', 'external': 'External',
    }

    prefix = ''
    if optional_match:
        prefix = '[OPTIONAL] '
    if manual_match:
        prefix = '[MANUAL DATE] '

    template_entries.append({
        'name': prefix + name,
        'phase': phase_labels.get(phase, phase),
        'owner': owner_labels.get(owner, owner),
        'lead_time': lead_time,
        'deps': deps,
        'notes': notes,
    })

for row_idx, entry in enumerate(template_entries, 2):
    values = [
        entry['name'], entry['phase'], entry['owner'], entry['lead_time'],
        entry['deps'], entry['notes'], '', '', ''
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = WRAP if col_idx in (1, 5, 6, 8, 9) else CENTER if col_idx == 4 else Alignment(vertical='top')
        cell.border = BORDER
        cell.font = Font(name='Arial', size=10)

    # Yellow the review columns
    for col in (7, 8, 9):
        ws2.cell(row=row_idx, column=col).fill = PatternFill('solid', fgColor='FFF9C4')

# ════════════════════════════════════════════════════════════════════
# TAB 3: LAUNCH CONFIG
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Launch Configs')

config_headers = [
    'Launch Name', 'DTC Launch Date', 'Sephora Launch Date', 'Amazon Launch Date',
    'Tier', 'Type', 'Content Production', 'Category', 'Status',
    'REVIEW — Correct Dates', 'REVIEW — Other Notes'
]
config_widths = [32, 16, 18, 18, 8, 16, 18, 12, 14, 24, 40]

for col_idx, (header, width) in enumerate(zip(config_headers, config_widths), 1):
    cell = ws3.cell(row=1, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws3.column_dimensions[get_column_letter(col_idx)].width = width

ws3.freeze_panes = 'A2'

type_labels = {
    'new_product': 'New Product', 'product_extension': 'Product Extension',
    'campaign': 'Campaign', 'seasonal': 'Seasonal', 'collaboration': 'Collaboration',
}
cpt_labels = {
    'none': 'None', 'no_tech': 'No Tech', 'with_tech': 'With Tech',
    'landing_page': 'Landing Page',
}
status_labels = {
    'planning': 'Planning', 'in_progress': 'In Progress',
    'launched': 'Launched', 'post_launch': 'Post-Launch', 'archived': 'Archived',
}

for row_idx, launch in enumerate(sorted(launches, key=lambda l: l.get('launchDate', '')), 2):
    values = [
        launch['name'],
        launch.get('launchDate', ''),
        launch.get('sephoraLaunchDate', ''),
        launch.get('amazonLaunchDate', ''),
        launch.get('tier', ''),
        type_labels.get(launch.get('launchType', ''), launch.get('launchType', '')),
        cpt_labels.get(launch.get('contentProductionType', ''), ''),
        launch.get('productCategory', ''),
        status_labels.get(launch.get('status', ''), launch.get('status', '')),
        '', '',
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = CENTER if col_idx in (2, 3, 4, 5, 9) else Alignment(vertical='top')
        cell.border = BORDER
        cell.font = Font(name='Arial', size=10)

    for col in (10, 11):
        ws3.cell(row=row_idx, column=col).fill = PatternFill('solid', fgColor='FFF9C4')

# ── Save ────────────────────────────────────────────────────────────
output_path = '/Users/juliaford/Desktop/GTM_Planner_Review.xlsx'
wb.save(output_path)
print(f"\nExported to: {output_path}")
print(f"  Tab 1: All Tasks — {row - 2} rows across {len([l for l in launches if l.get('status') != 'archived'])} launches")
print(f"  Tab 2: Task Template — {len(template_entries)} template tasks")
print(f"  Tab 3: Launch Configs — {len(launches)} launches")
print(f"\nYellow columns are for your review notes!")
